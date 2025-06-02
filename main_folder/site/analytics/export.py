# analytics/export.py
import io
import matplotlib.pyplot as plt
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook
from io import BytesIO
from PIL import Image as PILImage

def generate_price_dynamics_chart(prediction_result):
    """Генерация графика динамики цен"""
    plt.figure(figsize=(10, 6))
    
    # Получение данных из результата прогноза
    data = prediction_result.price_dynamics_data
    dates = data['dates']
    prices = data['prices']
    
    # Создание графика
    plt.plot(dates, prices, marker='o', linestyle='-', color='#1E88E5')
    plt.title('Прогноз динамики цен')
    plt.xlabel('Дата')
    plt.ylabel('Стоимость, руб.')
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    # Сохранение графика в байтовый поток
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=300)
    plt.close()
    
    buffer.seek(0)
    return buffer

def generate_comparison_chart(prediction_result):
    """Генерация графика сравнения инвестиций"""
    plt.figure(figsize=(10, 6))
    
    # Получение данных из результата прогноза
    data = prediction_result.comparison_data
    years = data['years']
    
    # Создание графика
    plt.plot(years, data['property'], marker='o', linestyle='-', label='Недвижимость', color='#1E88E5')
    plt.plot(years, data['stocks'], marker='s', linestyle='-', label='Акции', color='#FFC107')
    plt.plot(years, data['bonds'], marker='^', linestyle='-', label='Облигации', color='#4CAF50')
    plt.plot(years, data['deposits'], marker='x', linestyle='-', label='Депозиты', color='#F44336')
    
    plt.title('Сравнение доходности различных инвестиций')
    plt.xlabel('Год')
    plt.ylabel('Стоимость, руб.')
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.legend()
    plt.tight_layout()
    
    # Сохранение графика в байтовый поток
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=300)
    plt.close()
    
    buffer.seek(0)
    return buffer

def generate_factors_chart(prediction_result):
    """Генерация графика факторов влияния"""
    plt.figure(figsize=(10, 6))
    
    # Получение данных из результата прогноза
    factors = {
        'Привлекательность района': prediction_result.location_attractiveness_factor,
        'Транспортная доступность': prediction_result.transport_accessibility_factor,
        'Социальная инфраструктура': prediction_result.social_infrastructure_factor,
        'Перспективы развития': prediction_result.location_development_factor,
        'Макроэкономические факторы': prediction_result.macroeconomic_factor
    }
    
    # Создание графика
    plt.pie(
        factors.values(),
        labels=factors.keys(),
        autopct='%1.1f%%',
        startangle=90,
        colors=['#1E88E5', '#FFC107', '#4CAF50', '#F44336', '#9C27B0']
    )
    plt.axis('equal')
    plt.title('Факторы, влияющие на стоимость')
    plt.tight_layout()
    
    # Сохранение графика в байтовый поток
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=300)
    plt.close()
    
    buffer.seek(0)
    return buffer

def generate_pdf_report(prediction_request):
    """Генерация PDF отчета с результатами прогноза"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    
    # Получение результатов прогноза
    prediction_results = prediction_request.results.all()
    
    # Стили для документа
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']
    
    # Содержимое документа
    content = []
    
    # Заголовок
    content.append(Paragraph("Отчет по анализу недвижимости", title_style))
    content.append(Spacer(1, 12))
    
    # Информация об объекте
    property_data = prediction_request.property_data
    content.append(Paragraph("Информация об объекте", subtitle_style))
    content.append(Spacer(1, 6))
    
    property_info = [
        ["Параметр", "Значение"],
        ["Адрес", property_data.location_address],
        ["Площадь", f"{property_data.area} м²"],
        ["Тип помещения", dict(property_data.PROPERTY_TYPES)[property_data.property_type]],
        ["Год постройки", str(property_data.build_year)],
        ["Отделка", dict(property_data.FINISHING_TYPES)[property_data.finishing_type]],
        ["Этаж", f"{property_data.floor} из {property_data.total_floors}"]
    ]
    
    property_table = Table(property_info, colWidths=[200, 250])
    property_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.black),
        ('ALIGN', (0, 0), (1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (1, 0), 12),
        ('BACKGROUND', (0, 1), (1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    content.append(property_table)
    content.append(Spacer(1, 12))
    
    # Макроэкономические параметры
    content.append(Paragraph("Макроэкономические параметры", subtitle_style))
    content.append(Spacer(1, 6))
    
    macro_info = [
        ["Параметр", "Значение"],
        ["Инфляция", f"{prediction_request.inflation_rate}%"],
        ["Ставка ЦБ", f"{prediction_request.central_bank_rate}%"],
        ["Индекс потребительских цен", f"{prediction_request.consumer_price_index}%"],
        ["Темп роста ВВП", f"{prediction_request.gdp_growth_rate}%"],
        ["Ставка по ипотеке", f"{prediction_request.mortgage_rate}%"],
        ["Доходность депозитов", f"{prediction_request.deposit_rate}%"]
    ]
    
    macro_table = Table(macro_info, colWidths=[200, 250])
    macro_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.black),
        ('ALIGN', (0, 0), (1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (1, 0), 12),
        ('BACKGROUND', (0, 1), (1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    content.append(macro_table)
    content.append(Spacer(1, 24))
    
    # Результаты прогноза для каждого сценария
    for result in prediction_results:
        content.append(Paragraph(f"Сценарий: {dict(result.SCENARIO_TYPES)[result.scenario_type]}", subtitle_style))
        content.append(Spacer(1, 6))
        
        content.append(Paragraph(f"Прогнозируемая стоимость: {result.predicted_price:,.2f} руб.", normal_style))
        content.append(Paragraph(f"Ожидаемая доходность: {result.annual_yield}% в год", normal_style))
        content.append(Paragraph(f"Инвестиционный горизонт: {result.investment_horizon} лет", normal_style))
        content.append(Spacer(1, 12))
        
        # Добавление графиков
        # График динамики цен
        price_dynamics_img = generate_price_dynamics_chart(result)
        img = PILImage.open(price_dynamics_img)
        img = img.convert('RGB')
        img_temp = BytesIO()
        img.save(img_temp, 'JPEG')
        img_temp.seek(0)
        
        content.append(Paragraph("Динамика цен", normal_style))
        content.append(Image(img_temp, width=450, height=270))
        content.append(Spacer(1, 12))
        
        # График сравнения инвестиций
        comparison_img = generate_comparison_chart(result)
        img = PILImage.open(comparison_img)
        img = img.convert('RGB')
        img_temp = BytesIO()
        img.save(img_temp, 'JPEG')
        img_temp.seek(0)
        
        content.append(Paragraph("Сравнение доходности инвестиций", normal_style))
        content.append(Image(img_temp, width=450, height=270))
        content.append(Spacer(1, 12))
        
        # График факторов влияния
        factors_img = generate_factors_chart(result)
        img = PILImage.open(factors_img)
        img = img.convert('RGB')
        img_temp = BytesIO()
        img.save(img_temp, 'JPEG')
        img_temp.seek(0)
        
        content.append(Paragraph("Факторы влияния", normal_style))
        content.append(Image(img_temp, width=450, height=270))
        content.append(Spacer(1, 12))
        
        # Таблица с факторами влияния
        factors_info = [
            ["Фактор", "Влияние, %"],
            ["Привлекательность района (УРЖ)", f"{result.location_attractiveness_factor}%"],
            ["Транспортная доступность", f"{result.transport_accessibility_factor}%"],
            ["Социальная инфраструктура", f"{result.social_infrastructure_factor}%"],
            ["Перспективы развития локации", f"{result.location_development_factor}%"],
            ["Макроэкономические факторы", f"{result.macroeconomic_factor}%"]
        ]
        
        factors_table = Table(factors_info, colWidths=[250, 200])
        factors_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (1, 0), colors.black),
            ('ALIGN', (0, 0), (1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (1, 0), 12),
            ('BACKGROUND', (0, 1), (1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        content.append(factors_table)
        content.append(Spacer(1, 24))
    
    # Формирование документа
    doc.build(content)
    
    buffer.seek(0)
    return buffer

def generate_excel_report(prediction_request):
    """Генерация Excel отчета с результатами прогноза"""
    wb = Workbook()
    
    # Получение результатов прогноза
    prediction_results = prediction_request.results.all()
    property_data = prediction_request.property_data
    
    # Информация об объекте
    ws_property = wb.active
    ws_property.title = "Объект недвижимости"
    
    # Заголовки
    ws_property['A1'] = "Параметр"
    ws_property['B1'] = "Значение"
    
    # Данные об объекте
    property_info = [
        ["Адрес", property_data.location_address],
        ["Площадь", f"{property_data.area} м²"],
        ["Тип помещения", dict(property_data.PROPERTY_TYPES)[property_data.property_type]],
        ["Год постройки", str(property_data.build_year)],
        ["Отделка", dict(property_data.FINISHING_TYPES)[property_data.finishing_type]],
        ["Этаж", f"{property_data.floor} из {property_data.total_floors}"]
    ]
    
    for i, (param, value) in enumerate(property_info, start=2):
        ws_property[f'A{i}'] = param
        ws_property[f'B{i}'] = value
    
    # Макроэкономические параметры
    ws_property['D1'] = "Макроэкономический параметр"
    ws_property['E1'] = "Значение"
    
    macro_info = [
        ["Инфляция", f"{prediction_request.inflation_rate}%"],
        ["Ставка ЦБ", f"{prediction_request.central_bank_rate}%"],
        ["Индекс потребительских цен", f"{prediction_request.consumer_price_index}%"],
        ["Темп роста ВВП", f"{prediction_request.gdp_growth_rate}%"],
        ["Ставка по ипотеке", f"{prediction_request.mortgage_rate}%"],
        ["Доходность депозитов", f"{prediction_request.deposit_rate}%"]
    ]
    
    for i, (param, value) in enumerate(macro_info, start=2):
        ws_property[f'D{i}'] = param
        ws_property[f'E{i}'] = value
    
    # Результаты для каждого сценария
    for result in prediction_results:
        scenario_name = dict(result.SCENARIO_TYPES)[result.scenario_type]
        ws_scenario = wb.create_sheet(title=scenario_name)
        
        # Основная информация
        ws_scenario['A1'] = "Параметр"
        ws_scenario['B1'] = "Значение"
        
        scenario_info = [
            ["Прогнозируемая стоимость", f"{result.predicted_price:,.2f} руб."],
            ["Ожидаемая доходность", f"{result.annual_yield}% в год"],
            ["Инвестиционный горизонт", f"{result.investment_horizon} лет"]
        ]
        
        for i, (param, value) in enumerate(scenario_info, start=2):
            ws_scenario[f'A{i}'] = param
            ws_scenario[f'B{i}'] = value
        
        # Факторы влияния
        ws_scenario['A6'] = "Фактор влияния"
        ws_scenario['B6'] = "Вес, %"
        
        factors_info = [
            ["Привлекательность района (УРЖ)", result.location_attractiveness_factor],
            ["Транспортная доступность", result.transport_accessibility_factor],
            ["Социальная инфраструктура", result.social_infrastructure_factor],
            ["Перспективы развития локации", result.location_development_factor],
            ["Макроэкономические факторы", result.macroeconomic_factor]
        ]
        
        for i, (factor, weight) in enumerate(factors_info, start=7):
            ws_scenario[f'A{i}'] = factor
            ws_scenario[f'B{i}'] = weight
        
        # Динамика цен
        ws_dynamics = wb.create_sheet(title=f"{scenario_name} - Динамика цен")
        
        ws_dynamics['A1'] = "Дата"
        ws_dynamics['B1'] = "Прогнозируемая стоимость, руб."
        
        price_dynamics = result.price_dynamics_data
        for i, (date, price) in enumerate(zip(price_dynamics['dates'], price_dynamics['prices']), start=2):
            ws_dynamics[f'A{i}'] = date
            ws_dynamics[f'B{i}'] = price
        
        # Сравнение инвестиций
        ws_comparison = wb.create_sheet(title=f"{scenario_name} - Сравнение инвестиций")
        
        ws_comparison['A1'] = "Год"
        ws_comparison['B1'] = "Недвижимость, руб."
        ws_comparison['C1'] = "Акции, руб."
        ws_comparison['D1'] = "Облигации, руб."
        ws_comparison['E1'] = "Депозиты, руб."
        
        comparison_data = result.comparison_data
        for i, year in enumerate(comparison_data['years'], start=2):
            ws_comparison[f'A{i}'] = year
            ws_comparison[f'B{i}'] = comparison_data['property'][i-2]
            ws_comparison[f'C{i}'] = comparison_data['stocks'][i-2]
            ws_comparison[f'D{i}'] = comparison_data['bonds'][i-2]
            ws_comparison[f'E{i}'] = comparison_data['deposits'][i-2]
    
    # Сохранение в байтовый поток
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
